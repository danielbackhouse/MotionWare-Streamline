""" Title: DayDataAnalysis
    Purpose: Uses the Lights Out Index List, the Got Up Index List, and the
    activity, dates and times list of a single participant to create and
    calculate the waking hours analysis which is saved onto an excel document,
    one sheet per day, one workbook per participant.
    Author: Alan Yan and Daniel Backhouse
"""
#import extension libraries
from openpyxl import Workbook

#TODO: Allow user to change this!
filePath = r'C:\Users\dbackhou\Desktop\Day Data'
def findDayInfo(participant, LOIndexList, GUIndexList, activityList, datesList, timesList):
    """Finds the wake point for one sleep period
    
    Gets called from Study.py in the get_in_bed_times_noDiary function, utilizing
    the information found by that function. This function goes through each day
    by looking at the got up index and the corresponding light out index. Activity
    counts are then looked at, and from that, it is determined based on thresholds
    whether the participant is considered either sedentary, light or MVPA in movement
    Each day gets a sheet and each participant gets a file
    
    
    :param: The participant, the lights out and got up index lists and the 
    activity, dates and times lists for that participant
    :return: Returns nothing, saved as an excel document.
    :rtype: None
    """
    patientWorkbook = Workbook()
    daySheet = patientWorkbook.active
    errors = 0
    warnings = 0
    for index in range(len(LOIndexList)-1):
        if index:
            daySheet = patientWorkbook.create_sheet("Day " + str(index+1))
        if GUIndexList[index] >= LOIndexList[index+1]:
            print('Participant:', participant)
            print('Day:', index+1)
            print(GUIndexList[index], LOIndexList[index+1],)
            daySheet['A1'] = 'ERROR'
            daySheet.title = "Day " + str(index+1) + ' ERROR'
            errors += 1
            print("Errors:", errors)
        else:
            if LOIndexList[index+1]-GUIndexList[index] > 2160:
                daySheet.title = 'Day ' + str(index+1) + ' WARNING'
                daySheet['F1'] = 'WARNING: WAKING HOURS EXCEEDS 36 HOURS, LIKELY ERROR'
                warnings += 1
            else:      
                daySheet.title = "Day " + str(index+1)
            daySheet['A1'] = 'Participant ' + participant
            daySheet['A2'] = 'Time'
            daySheet['B2'] = 'Counts'
            daySheet['C2'] = 'Sedentary'
            daySheet['D2'] = 'Light'
            daySheet['E2'] = 'MVPA'
            daySheet['G2'] = 'TOTALS'
            daySheet['H2'] = 'SED'
            daySheet['I2'] = 'Light'
            daySheet['J2'] = 'MVPA'
            trimmedActivity = activityList[GUIndexList[index]: LOIndexList[index+1]]
            trimmedDates = datesList[GUIndexList[index]: LOIndexList[index+1]]
            trimmedTimes = timesList[GUIndexList[index]: LOIndexList[index+1]]
            daySheet['B1'] = 'Start Date'
            daySheet['C1'] = trimmedDates[0]
            daySheet['D1'] = 'End Date'
            daySheet['E1'] = trimmedDates[-1]
            sedentaryTotal = lightTotal = MVPATotal = 0
            for x in range(len(trimmedActivity)):
                daySheet.cell(row = x+3, column = 1).value = trimmedTimes[x]
                daySheet.cell(row = x+3, column = 2).value = trimmedActivity[x]
                sedentary = light = MVPA = 0
                if trimmedActivity[x] > 562.5:
                    MVPATotal += 1
                    MVPA = 1
                elif trimmedActivity[x] < 172.5:
                    sedentaryTotal += 1
                    sedentary = 1
                else:
                    lightTotal += 1
                    light = 1
                daySheet.cell(row = x+3, column = 3).value = sedentary
                daySheet.cell(row = x+3, column = 4).value = light
                daySheet.cell(row = x+3, column = 5).value = MVPA
            daySheet['H3'] = sedentaryTotal
            daySheet['I3'] = lightTotal
            daySheet['J3'] = MVPATotal
            letters= ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
            for x in range(10):
                daySheet.column_dimensions[letters[x]].width = 12.0
            daySheet.column_dimensions['A'].width = 15.0
    if errors:
        error = ' Errors, ' + str(errors)
    else:
        error = ''
    if warnings:
        warning  = ', Warnings, ' + str(warnings)
    else:
        warning = ''
    patientWorkbook.save(filePath + '\BT-'+ participant + error + warning + '.xlsx')
    return